# import xlwings
# from xlwings.constants import AutoFillType

# autofill_dic = {}

# column_sc = dataframe_mapper['Source Coordinates']
# column_dc = dataframe_mapper['Destination Coordinates']
# cell_value = [column_sc + str(input_file_ws.max_row)]
# fill_max_columns = [column_dc + ':' + column_dc[:-1] + str(input_file_ws.max_row)]

# for i in range(len(cell_value)):
#     autofill_dic = {cell_value[i]: fill_max_columns[i]}

# for range_dic in autofill_dic:
#     input_xw_sheet.range(range_dic).api.AutoFill(input_xw_sheet.range(autofill_dic[range_dic]).api, AutoFillType.xlFillDefault)

import string
from openpyxl.utils import get_column_letter
def col2num(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num

a = col2num('S')

b = get_column_letter(1)
print(b)