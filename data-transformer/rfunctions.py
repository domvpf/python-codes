from xlwings.constants import AutoFillType
from openpyxl.utils import get_column_letter
import string

def call_retrieve_functions(rfunction_name, dataframe_mapper, input_file_ws, input_xw_sheet, output_xw_sheet, config_openpyxl_ws, process_row_start):
    rfunction_name = rfunction_name.strip()
    
    if rfunction_name == 'R Function 2':
        rfunction_2(dataframe_mapper, input_file_ws, input_xw_sheet, config_openpyxl_ws, process_row_start)
    if rfunction_name == 'R Function 4':
        if dataframe_mapper['Plot to Where'] == 'Input':
            rfunction_4_input(dataframe_mapper, input_file_ws, input_xw_sheet, config_openpyxl_ws, process_row_start)
        if dataframe_mapper['Plot to Where'] == 'Output':
            rfunction_4_output(dataframe_mapper, input_file_ws, input_xw_sheet, output_xw_sheet, config_openpyxl_ws, process_row_start)
    if rfunction_name == 'R Function 12':
        rfunctions_12(dataframe_mapper, input_file_ws, input_xw_sheet)
    if rfunction_name == 'R Function 13':
        rfunctions_13(dataframe_mapper,input_file_ws, input_xw_sheet, config_openpyxl_ws, process_row_start)
    if rfunction_name == 'R Function 16':
        rfunction_16(dataframe_mapper, input_xw_sheet)
    if rfunction_name == 'R Function 17':
        rfunction_17(dataframe_mapper, input_file_ws, input_xw_sheet, config_openpyxl_ws, process_row_start)
    
def rfunction_2(dataframe_mapper, input_file_ws, input_xw_sheet, config_openpyxl_ws, process_row_start):
    row_list = []
    values = config_openpyxl_ws[process_row_start+2]
    values = [values[x].value for x in range(len(values))]
    row_list.append(values)

    for config_row_data in row_list:
        config_column = [[config_row_data[1],config_row_data[2]]]

        for data in config_column:
            column_range, row_start = data
            cells, columns = column_range.split(':')
            fill_source_coordinates =  [cells + ':' + columns + str(input_file_ws.max_row)]
            fill_destination_coordinates = [dataframe_mapper['Destination Coordinates'] + ':' + dataframe_mapper['Destination Coordinates'][:-1] + str(input_file_ws.max_row)]
            
            for i in range(len(fill_source_coordinates)):
                input_xw_sheet.range(fill_source_coordinates[i]).value = input_xw_sheet.range(fill_destination_coordinates[i]).value 

def rfunction_16(dataframe_mapper, input_xw_sheet):
    excel_formula = str(dataframe_mapper['Excel Formula'].strip())
    input_xw_sheet[dataframe_mapper['Destination Coordinates']].value = excel_formula

def rfunction_17(dataframe_mapper,input_file_ws, input_xw_sheet, config_openpyxl_ws, process_row_start):
    row_list = []
    values = config_openpyxl_ws[process_row_start+2]
    values = [values[x].value for x in range(len(values))]
    row_list.append(values)

    for config_row_data in row_list:
        config_column = [[config_row_data[1],config_row_data[13]]]

        for data in config_column:
            column_range, row_start = data
            columns, rows = column_range.split(':')
            cell_value = [columns + str(row_start) + ':' + rows + str(row_start)]
            fill_max_columns =  [columns + str(row_start) + ':' + rows + str(input_file_ws.max_row)]
            for i in range(len(cell_value)):
                autofill_dic = {cell_value[i]: fill_max_columns[i]}
            
            for range_dic in autofill_dic:
                input_xw_sheet.range(range_dic).api.AutoFill(input_xw_sheet.range(autofill_dic[range_dic]).api, AutoFillType.xlFillDefault)

def rfunction_4_input(dataframe_mapper, input_file_ws, input_xw_sheet, config_openpyxl_ws, process_row_start):
    row_list = []
    values = config_openpyxl_ws[process_row_start+2]
    values = [values[x].value for x in range(len(values))]
    row_list.append(values)

    for config_row_data in row_list:
        config_column = [[config_row_data[1],config_row_data[2]]]

        for data in config_column:
            column_range, row_start = data
            cells, columns = column_range.split(':')
            fill_source_coordinates =  [cells + ':' + columns + str(input_file_ws.max_row)]
            fill_destination_coordinates = [dataframe_mapper['Destination Coordinates'] + ':' + dataframe_mapper['Destination Coordinates'][:-1] + str(input_file_ws.max_row)]
            
            for i in range(len(fill_source_coordinates)):
                input_xw_sheet.range(fill_source_coordinates[i]).value = input_xw_sheet.range(fill_destination_coordinates[i]).value 

def rfunction_4_output(dataframe_mapper, input_file_ws, input_xw_sheet, output_xw_sheet, config_openpyxl_ws, process_row_start):
    row_list = []
    values = config_openpyxl_ws[process_row_start+2]
    values = [values[x].value for x in range(len(values))]
    row_list.append(values)

    for config_row_data in row_list:
            config_column = [[config_row_data[1],config_row_data[2]]]

            for data in config_column:
                column_range, row_start = data
                cells, columns = column_range.split(':')

                # DOING SOME LITTLE MATHS
                first_column = cells[:-1]
                cell_to_index = column_to_index(first_column)
                columns_to_index = column_to_index(columns)

                subtracted_column = columns_to_index - cell_to_index
                final_column = get_column_letter(subtracted_column+1)

                subtract_columns_to_match = int(dataframe_mapper['Destination Coordinates'][-1]) - int(cells[-1])

                fill_source_coordinates =  [cells + ':' + columns + str(input_file_ws.max_row)]
                fill_destination_coordinates = [dataframe_mapper['Destination Coordinates'] + ':' + final_column + str(input_file_ws.max_row + 7)]

                for i in range(len(fill_source_coordinates)):
                    output_xw_sheet.range(fill_destination_coordinates[i]).value = input_xw_sheet.range(fill_source_coordinates[i]).value

def rfunctions_12(dataframe_mapper, input_file_ws, input_xw_sheet):
    startswith_list = []
    not_startswith_list = []

    fill_max_columns =  str(input_file_ws.max_row)
    cell = dataframe_mapper['Source Coordinates'] + fill_max_columns
    cell_range = input_xw_sheet.range(cell).value
    get_last_cell = input_xw_sheet.cells.end('down').end('right').address.replace('$', '')

    if dataframe_mapper['String Condition'] == 'startsWith':
        for index, value in enumerate(cell_range):
            if value.startswith(str(dataframe_mapper['Condition Value']).replace('.0', '')):
                startswith_list.append(index)
                # cell = 'A' + str(index+2) + ':' + get_last_cell[0] + str(index+2)
                # start_cell = 'A' + str(index+2)
                # end_cell = get_last_cell[0] + str(index+2)
            else:
                not_startswith_list.append(index)
                cell = 'A' + str(index+2) + ':' + get_last_cell[0] + str(index+2)
                input_xw_sheet.range(cell).value = None

def rfunctions_13(dataframe_mapper,input_file_ws, input_xw_sheet, config_openpyxl_ws, process_row_start):
    fix_excel_formula = dataframe_mapper['Excel Formula'].replace('[]', str(dataframe_mapper['Row Start']))
    excel_formula = str(fix_excel_formula.strip().replace('.0', ''))
    input_xw_sheet[dataframe_mapper['Destination Coordinates']].value = excel_formula

    row_list = []
    values = config_openpyxl_ws[process_row_start+2]
    values = [values[x].value for x in range(len(values))]
    row_list.append(values)

    for config_row_data in row_list:
        config_column = [[config_row_data[2],config_row_data[13],excel_formula]]

        for data in config_column:
            # column_range, row_start = data
            # columns, rows = column_range.split(':')
            cell_value = [data[0][0] + str(data[1]) + ':' + data[0][0] + str(data[1])]
            fill_max_columns =  [data[0][0] + str(data[1])+ ':' + data[0][0] + str(input_file_ws.max_row)]

            for i in range(len(cell_value)):
                autofill_dic = {cell_value[i]: fill_max_columns[i]}
            
            for range_dic in autofill_dic:
                input_xw_sheet.range(range_dic).api.AutoFill(input_xw_sheet.range(autofill_dic[range_dic]).api, AutoFillType.xlFillDefault)

def column_to_index(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num