import openpyxl, xlwings, argparse, time, traceback, datetime
import pandas as pd
from xlwings.constants import AutoFillType

def autofill_process(config_path, config_worksheet_name):

    config = openpyxl.load_workbook(config_path)
    config_ws = config[config_worksheet_name]

    row_list = []
    autofill_dic = {}
    app = xlwings.App(visible = False)

    for rows in range(2, config_ws.max_row + 1):
        values = config_ws[rows]
        values = [values[x].value for x in range(len(values))]
        row_list.append(values)

        # input_file = openpyxl.load_workbook(values[0])
        # input_ws = input_file[values[2]]

        # input_xw = xlwings.Book(values[0])
        # input_xw_sheet = input_xw.sheets(values[2])
    for config_row_data in row_list:
        input_file = openpyxl.load_workbook(config_row_data[0])
        input_ws = input_file[config_row_data[2]]

        input_xw = xlwings.Book(config_row_data[0])
        input_xw_sheet = input_xw.sheets(config_row_data[2])
        config_column = [[config_row_data[-2],config_row_data[-1]]]
        
        for data in config_column:
            column_range, row_start = data
            columns, rows = column_range.split(':')
            print(column_range)
            cell_value = [columns + str(row_start) + ':' + rows + str(row_start)]
            fill_max_columns =  [columns + str(row_start) + ':' + rows + str(input_ws.max_row)]

            for i in range(len(cell_value)):
                autofill_dic = {cell_value[i]: fill_max_columns[i]}
                print(autofill_dic)
            for range_dic in autofill_dic:
                input_xw_sheet.range(range_dic).api.AutoFill(input_xw_sheet.range(autofill_dic[range_dic]).api, AutoFillType.xlFillDefault)
                input_xw.save()
    input_xw.app.quit()
    app.kill()
def main(args=None):
    parser = argparse.ArgumentParser()
    parser.add_argument("config_path")
    parser.add_argument("config_worksheet_name")
    
    args =  parser.parse_args()

    dt_string = datetime.datetime.now().strftime("%m-%d-%Y %H-%M-%S")
    log_file = dt_string + ' - ' + 'log_file.log'

    try:
        autofill_process(args.config_path, args.config_worksheet_name)

    except Exception as exception:
        traceback.print_exception(
            type(exception),
            exception,
            exception.__traceback__
        )
        with open(log_file, "w") as file:
            file.write(f"Error in file: {exception}")

if __name__ == '__main__':
    start_time = time.time()
    main()
    print("--- %s seconds ---" % (time.time() - start_time))