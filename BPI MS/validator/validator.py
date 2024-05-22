from rules import Rules, Template

from xlrd import open_workbook, XLRDError
from openpyxl import load_workbook, Workbook
from datetime import datetime
from collections import OrderedDict
import numpy as np
import pandas as pd
import argparse
import csv
import os
import traceback
import re
import operator
import time
import itertools
import sys

trailer_count = 0
trailer_hash = 0

def get_file_name(dsf):
    [file_path, file_extension] = os.path.splitext(dsf)
    file_name = os.path.basename(file_path)
    return file_name

def write_to_csv(output_contents, output_path):

    if output_contents:
        with open(output_path, "w", newline="", encoding="utf-8") as file:
            w = csv.DictWriter(file, fieldnames=output_contents[0].keys())
            w.writeheader()
            w.writerows(output_contents)

def read_csv(dsf,dir_name, file_ext, dsf_type):
    records = []
    headers_new = []
    if file_ext=='xlsx':
        read_file = pd.read_excel(dsf)
        file_name_csv = get_file_name(dsf) + '.csv'

        dir_name = dir_name + "\\" + file_name_csv

        for col in read_file.columns:

            if re.search("[.][1-9]", col):
                counter = col.split(".")[-1]
                counter = int(counter) + 1
                col = col.split(".")
                col = ''.join(col)[:-1] + " " + str(counter)
                headers_new.append(col)
            else:
                headers_new.append(col)


        read_file.to_csv(dir_name, index = None, header=headers_new, float_format='%.2f')

        with open(dir_name) as f:
            reader = csv.DictReader(f)
            for line in reader:
                records.append(line)

            header = records[0].keys()

        return header, records

    elif file_ext=='csv' and dsf_type==4:

        if Rules.check_csv(dsf):# File level validation# 1
            dir_name = dir_name + ".csv"
            print(dir_name)

            if Rules.check_csv(dsf):# File level validation# 1
                with open(dir_name) as file:
                    reader_osp_headers = csv.reader(file)
                    headers_osp = next(reader_osp_headers)
                    header_hash = headers_osp[1]
                    # print("Header hash: ", header_hash)
                    headers_osp = next(reader_osp_headers)
                    headers_osp = next(reader_osp_headers)
                    header = headers_osp
                    if len(headers_osp) != len(set(headers_osp)):
                        raise Exception('Duplicate Headers')

                file.close()

                try:
                    with open(dsf) as f:
                        read = csv.reader(f)
                        # header = next(read)
                        next(read)
                        next(read)

                        reader = csv.DictReader(f)
                        for line in reader:
                            records.append(line)

                    # print(records)
                except UnicodeDecodeError:
                   raise Exception("Special character detected!")
                return header, records, header_hash
    else:
        headers_osp = []

        if Rules.check_csv(dsf):# File level validation# 1
            dir_name = dir_name + ".csv"
            print(dir_name)
            try:
                with open(dir_name) as file:
                    reader_osp_headers = csv.reader(file)
                    headers_osp = next(reader_osp_headers)
                    if len(headers_osp) != len(set(headers_osp)):
                        raise Exception('Duplicate Headers')
                file.close()
            except UnicodeDecodeError:
               raise Exception("Special character detected!")
            try:
                with open(dir_name) as f:
                    reader = csv.DictReader(f)

                    for line in reader:
                        records.append(line)

                    header = records[0].keys()
            except UnicodeDecodeError:
               raise Exception("Special character detected!")

            return header, records
        else:
            raise Exception('Not a csv')


def dsf_validator(dsf, dev_worksheet, comp_template, dsf_type, dir_name, file_ext):
    print(dev_worksheet)
    workbook = open_workbook(dev_worksheet)
    content = workbook.sheet_by_name('DSF Fields')
    headers = content.row_values(0)[1:-1]

    row = []

    for i in range(2, content.nrows):
        temp = {}
        for j in range(0, len(headers)):
            temp[headers[j]] = str(content.row_values(i)[1:][j])
        row.append(temp)
    dsf_fields = row[:-12]

    headers_dsf = []
    for d in dsf_fields:
        if 'DSF FIELDS <START>' in d:
            headers_dsf.append(d['DSF FIELDS <START>'])

    cont = load_workbook(dev_worksheet)
    ws = cont["Project Info"]

    data = []
    for row in ws.rows:
        for cell in row:
            if cell.value != None:
                data.append(cell.value)

    cont.close()

    rules = []
    for i,w in enumerate(data):
        try:
            if w == 'File Validation Rules:':
                x = data[i+1:]
                for i in x:
                    rules.append(i)
        except:
            rules == ''

    rules_h = rules[:-6:3]
    rules_v = rules[1:-6:3]
    validation_rules = dict(zip(rules_h,rules_v))
    # print(validation_rules)

    if dsf_type == 4:
        header, records, header_hash = read_csv(dsf,dir_name, file_ext, dsf_type)
    else:
        header, records = read_csv(dsf,dir_name, file_ext, dsf_type)

    # if validation_rules['Validate file is CSV'] and not Rules.check_csv(dsf):# File level validation# 1
    #     raise Exception('Not a csv')

    if validation_rules['At least one record'] and not Rules.at_least_one_record(len(records)):# File level validation# 3
        raise Exception('Has no Records')

    missing = Rules.check_header_fields(headers_dsf, header)
    if validation_rules['Validate Header Fields'] and missing: # File level validation# 2
        raise Exception(f'Missing headers: {missing}')


    if dsf_type == 4:
        if validation_rules['Header record validation']:
            print("Header hash: ", header_hash)
            if not Rules.check_header_hash(os.path.basename(dsf), header_hash):
                raise Exception('Header Hash does not match')

        if validation_rules['Trailer record count validation'] or validation_rules['Trailer record Hash validation']:
            trailer = list(records.pop().values())
            trailer_count, trailer_hash = trailer[1:3]
            print("Trailer count: ",trailer_count)
            print("Trailer hash: ",trailer_hash)
            # # if trailer_hash == "":
            # #     trailer_hash = trailer[3]
            # #     matches = (x for x in trailer if len(x) > 1)
            # #     for i in matches:
            # #         print(i)
            # #     print("Trailer hash: ",trailer_hash)
            # # else:
            #
            # print(trailer)
            if not Rules.trailer_hash_exists(trailer):
                raise Exception('Trailer Hash does not exists')
            if validation_rules['Trailer record count validation'] and not Rules.check_trailer_count(records, int(trailer_count)):# File level validation# 5
                raise Exception('Trailer count does not match')
            if validation_rules['Trailer record Hash validation'] and not Rules. check_trailer_hash(records, int(trailer_hash)):# File level validation# 5
                raise Exception('Trailer Hash does not match')

    if validation_rules['Contract type validation']:
        df_contract_types = pd.read_excel(dev_worksheet, sheet_name='Contract Types', usecols="A")
        dict_contract_types = df_contract_types.values.tolist()
        valid_contract_types = []
        for i in dict_contract_types:
            valid_contract_types.append(i[0])

        contract_type_field = valid_contract_types.pop(0)
        valid_contract_types = [each_string.lower() for each_string in valid_contract_types]
        # print(valid_contract_types)

    mandatory_p = []
    mandatory_c = []
    DSF_fields_p = {}
    DSF_fields_c = {}

    for d in dsf_fields:
        # DSF_fields[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'], 'Value':d['Consistent Default Value']}
        if dsf_type == 1:
            if d['Field Attribute'] == 'P':
                mandatory_p.append(d['DSF FIELDS <START>'])
                DSF_fields_p[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
            elif d['Field Attribute'] == 'C':
                mandatory_c.append(d['DSF FIELDS <START>'])
                DSF_fields_c[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
            elif d['Field Attribute'] =='B':
                mandatory_p.append(d['DSF FIELDS <START>'])
                mandatory_c.append(d['DSF FIELDS <START>'])
                DSF_fields_c[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
                DSF_fields_p[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
            else:
                DSF_fields_c[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'],'Custom Rule':d['Custom Rule']}
                DSF_fields_p[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
        elif dsf_type == 0:
            if d['Field Attribute'] == 'P':
                mandatory_p.append(d['DSF FIELDS <START>'])
                DSF_fields_p[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
            elif d['Field Attribute'] == 'C':
                mandatory_c.append(d['DSF FIELDS <START>'])
                DSF_fields_c[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
            elif d['Field Attribute'] =='B':
                mandatory_p.append(d['DSF FIELDS <START>'])
                mandatory_c.append(d['DSF FIELDS <START>'])
                DSF_fields_c[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
                DSF_fields_p[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
            else:
                DSF_fields_c[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'],'Custom Rule':d['Custom Rule']}
                DSF_fields_p[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
        elif dsf_type == 2:
            if d['Field Attribute'] == 'P':
                mandatory_p.append(d['DSF FIELDS <START>'])
                DSF_fields_p[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
            elif d['Field Attribute'] == 'C':
                mandatory_c.append(d['DSF FIELDS <START>'])
                DSF_fields_c[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
            elif d['Field Attribute'] =='B':
                mandatory_p.append(d['DSF FIELDS <START>'])
                mandatory_c.append(d['DSF FIELDS <START>'])
                DSF_fields_c[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
                DSF_fields_p[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
            else:
                DSF_fields_c[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'],'Custom Rule':d['Custom Rule']}
                DSF_fields_p[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
        elif dsf_type == 3:
            if d['Field Attribute'] == 'P':
                mandatory_p.append(d['DSF FIELDS <START>'])
                DSF_fields_p[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
            elif d['Field Attribute'] == 'C':
                mandatory_c.append(d['DSF FIELDS <START>'])
                DSF_fields_c[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
            elif d['Field Attribute'] =='B':
                mandatory_p.append(d['DSF FIELDS <START>'])
                mandatory_c.append(d['DSF FIELDS <START>'])
                DSF_fields_c[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
                DSF_fields_p[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
            else:
                DSF_fields_c[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'],'Custom Rule':d['Custom Rule']}
                DSF_fields_p[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
        elif dsf_type == 4:
            if d['Field Attribute'] == 'P':
                mandatory_p.append(d['DSF FIELDS <START>'])
                DSF_fields_p[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
            elif d['Field Attribute'] == 'C':
                mandatory_c.append(d['DSF FIELDS <START>'])
                DSF_fields_c[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
            elif d['Field Attribute'] =='B':
                mandatory_p.append(d['DSF FIELDS <START>'])
                mandatory_c.append(d['DSF FIELDS <START>'])
                DSF_fields_c[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
                DSF_fields_p[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}
            else:
                DSF_fields_c[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'],'Custom Rule':d['Custom Rule']}
                DSF_fields_p[d['DSF FIELDS <START>']] = {'Length':d['Length'], 'Type':d['Type'],
                'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX'], 'Custom Rule':d['Custom Rule']}

    valid_tl = []
    invalid_tl = []
    others_tl = []
    mf_rows = []
    others_rows = []
    mf_h = []
    client_type = 'P'
    type_flag = 0
    length_flag = 0

    content_custom_rules = workbook.sheet_by_name('Custom Rules')
    headers_custom_rules = content_custom_rules.row_values(0)[0:-1]


    i=1
    custom_rules_fields = []
    custom_rules_operators = []
    custom_rules_values = []
    custom_rules_conditions = []
    while headers_custom_rules[i] != 'DSF Field':
        if headers_custom_rules[i] != 'DSF Field':
            if 'Field' in headers_custom_rules[i]:
                custom_rules_fields.append(headers_custom_rules[i])
            elif 'Operator' in headers_custom_rules[i]:
                custom_rules_operators.append(headers_custom_rules[i])
            elif 'Value' in headers_custom_rules[i]:
                custom_rules_values.append(headers_custom_rules[i])
            elif 'Condition' in headers_custom_rules[i]:
                custom_rules_conditions.append(headers_custom_rules[i])
        elif headers_custom_rules[i] == 'DSF Field':
            break
        i += 1

    custom_rules_headers = []

    row_custom_rules = []

    for i in range(1, content_custom_rules.nrows):
        temp = {}
        for j in range(0, len(headers_custom_rules)):
            temp[headers_custom_rules[j]] = str(content_custom_rules.row_values(i)[0:][j])
        row_custom_rules.append(temp)


    dsf_fields_custom_rules_temp = row_custom_rules[1:]

    headers_dsf_custom_rules = []

    for d in dsf_fields_custom_rules_temp:
        if 'Rule no' in d:
            headers_dsf_custom_rules.append(d['Rule no'])

    dsf_fields_custom_rules = {}
    dsf_fields_headers_to_edit_temp = []

    field_operator_value_condition={}

    for i in custom_rules_fields:
        for dict_custom_headers in dsf_fields_custom_rules_temp:
            for header in dict_custom_headers:
                if i == header:
                    field_operator_value_condition[i] = dict_custom_headers[header]

    for i in dsf_fields_custom_rules_temp:
        dsf_fields_custom_rules[i['Rule no']] = {'DSF Field': i['DSF Field'], 'Field Attribute':i['Field Attribute'],'Length':i['Length'],'Type':i['Type'],
        'Type':i['Type'],'Consistent Default Value':i['Consistent Default Value'], 'Min':i['MIN'], 'Max':i['MAX']}

    for j in custom_rules_fields:
        for dict_custom_headers in dsf_fields_custom_rules_temp:
            for header in dict_custom_headers:
                if j == header:
                    temp_dict_fields = {}
                    temp_dict_fields[j] = dict_custom_headers[header]
                    dsf_fields_custom_rules[dict_custom_headers['Rule no']].update(temp_dict_fields)
                    dsf_fields_headers_to_edit_temp.append(dict_custom_headers[header])

    for j in custom_rules_operators:
        for dict_custom_headers in dsf_fields_custom_rules_temp:
            for header in dict_custom_headers:
                if j == header:
                    temp_dict_fields = {}
                    temp_dict_fields[j] = dict_custom_headers[header]
                    dsf_fields_custom_rules[dict_custom_headers['Rule no']].update(temp_dict_fields)

    for j in custom_rules_values:
        for dict_custom_headers in dsf_fields_custom_rules_temp:
            for header in dict_custom_headers:
                if j == header:
                    temp_dict_fields = {}
                    temp_dict_fields[j] = dict_custom_headers[header]
                    dsf_fields_custom_rules[dict_custom_headers['Rule no']].update(temp_dict_fields)

    for j in custom_rules_conditions:
        for dict_custom_headers in dsf_fields_custom_rules_temp:
            for header in dict_custom_headers:
                if j == header:
                    temp_dict_fields = {}
                    temp_dict_fields[j] = dict_custom_headers[header]
                    dsf_fields_custom_rules[dict_custom_headers['Rule no']].update(temp_dict_fields)

    dsf_fields_headers_to_edit = set(dsf_fields_headers_to_edit_temp)
    dsf_fields_headers_to_edit = list(filter(None, dsf_fields_headers_to_edit))

    for header, req in DSF_fields_p.items():
        if req['Custom Rule'] == 'Yes':
            custom_rules_headers.append(header)

    edited_headers = []

    mandatory_edited_flag = False

    for record in records:

        mandatory_c_temp = list(mandatory_c)
        mandatory_p_temp = list(mandatory_p)

        if dsf_type>=0 and dsf_type <=4:
            if dsf_type==1:
                try:
                     policy_type = record['POLICY NUMBER'][:2]
                     policy_number_holder = record['POLICY NUMBER']
                     record['POLICY NUMBER'] = policy_type
                     print(record['POLICY NUMBER'], policy_number_holder)
                except KeyError:
                    raise Exception("DSF input file not compatible with DSF argument")
                except:
                  raise Exception("DSF input file not compatible with DSF argument")
            elif dsf_type==0:
                try:
                     print("\n",record['POLICY NO'])
                except KeyError:
                   print("\n",record['Policy No.'])
                except KeyError:
                   print("\n",record['Policy No'])
                # finally:
                #   raise Exception("DSF input file not compatible with DSF argument")
            elif dsf_type==4:
                try:
                     print("\n",record['Policy No.'])
                except KeyError:
                   raise Exception("DSF input file not compatible with DSF argument")
                except:
                  raise Exception("DSF input file not compatible with DSF argument")
            elif dsf_type==2:
                try:
                  if record['Client Type'].lower() == "personal":
                      client_type = 'P'
                      print("\n",record['Policy No']," - Personal")
                  elif record['Client Type'].lower() == "corporate":
                      client_type = 'C'
                      print("\n",record['Policy No']," - Corporate")
                except KeyError:
                  raise Exception("DSF input file not compatible with DSF argument")
                except:
                  raise Exception("DSF input file not compatible with DSF argument")
            elif dsf_type==3:
                try:
                  client_type = 'P'
                except KeyError:
                  raise Exception("DSF input file not compatible with DSF argument")
                except:
                  raise Exception("DSF input file not compatible with DSF argument")

            for header_with_custom_rule in custom_rules_headers: #custom rules

                if header_with_custom_rule in record:
                    print(header_with_custom_rule)
                    for key,value_dsf in dsf_fields_custom_rules.items():

                        count_col = 0
                        actual_counter = 0
                        field_counter = 1
                        dict_lists_to_check = []
                        for i,j in zip(custom_rules_fields,custom_rules_values):


                            if value_dsf[i] == 'Agent Code':
                                value_dsf[j] = value_dsf[j].split('.')[0]
                            count_col += 1
                            num_conditions = 0
                            field_counter = 0
                            actual_conditions = []

                            for m in custom_rules_fields:
                                if value_dsf[m]:
                                    field_counter += 1

                            for k in custom_rules_conditions:
                                if value_dsf[k]:
                                    num_conditions += 1
                                    actual_conditions.append(value_dsf[k])

                            if value_dsf[i]:
                                check = True
                                actual_counter += 1

                                if len(actual_conditions) == 0:

                                    try:
                                        actual = record[value_dsf[i]]
                                    except KeyError:
                                        pass

                                    if header_with_custom_rule == value_dsf[i]:


                                        if value_dsf[i] == 'Agent Code':
                                            value_dsf[j] = value_dsf[j].split('.')[0]
                                        if record[header_with_custom_rule] == value_dsf[j]:
                                            # print("value in custom rule")
                                            header_to_edit = value_dsf['DSF Field']
                                            if value_dsf['Field Attribute'] == 'P':
                                                if not header_to_edit in mandatory_p_temp:
                                                    mandatory_p_temp.append(header_to_edit)
                                                DSF_fields_p[header_to_edit] = {'Field Attribute':value_dsf['Field Attribute'],'Length':value_dsf['Length'],
                                                'Type':value_dsf['Type'],'Value':value_dsf['Consistent Default Value'], 'Min':value_dsf['Min'], 'Max':value_dsf['Max']}
                                            elif value_dsf['Field Attribute'] == 'C':
                                                if not header_to_edit in mandatory_c_temp:
                                                    mandatory_c_temp.append(header_to_edit)
                                                DSF_fields_c[header_to_edit] =  {'Field Attribute':value_dsf['Field Attribute'],'Length':value_dsf['Length'],
                                                'Type':value_dsf['Type'],'Value':value_dsf['Consistent Default Value'], 'Min':value_dsf['Min'], 'Max':value_dsf['Max']}
                                            elif value_dsf['Field Attribute'] =='B':
                                                if not header_to_edit in mandatory_p_temp:
                                                    mandatory_p_temp.append(header_to_edit)
                                                if not header_to_edit in mandatory_c_temp:
                                                    mandatory_c_temp.append(header_to_edit)
                                                DSF_fields_c[header_to_edit] =  {'Field Attribute':value_dsf['Field Attribute'],'Length':value_dsf['Length'],
                                                'Type':value_dsf['Type'],'Value':value_dsf['Consistent Default Value'], 'Min':value_dsf['Min'], 'Max':value_dsf['Max']}
                                                DSF_fields_p[header_to_edit] =  {'Field Attribute':value_dsf['Field Attribute'],'Length':value_dsf['Length'],
                                                'Type':value_dsf['Type'],'Value':value_dsf['Consistent Default Value'], 'Min':value_dsf['Min'], 'Max':value_dsf['Max']}
                                            else:
                                                if header_to_edit in mandatory_p_temp:
                                                    mandatory_p_temp.remove(header_to_edit)
                                                if header_to_edit in mandatory_c_temp:
                                                    mandatory_c_temp.remove(header_to_edit)
                                                DSF_fields_c[header_to_edit] =  {'Field Attribute':value_dsf['Field Attribute'],'Length':value_dsf['Length'],
                                                'Type':value_dsf['Type'],'Value':value_dsf['Consistent Default Value'], 'Min':value_dsf['Min'], 'Max':value_dsf['Max']}
                                                DSF_fields_p[header_to_edit] =  {'Field Attribute':value_dsf['Field Attribute'],'Length':value_dsf['Length'],
                                                'Type':value_dsf['Type'],'Value':value_dsf['Consistent Default Value'], 'Min':value_dsf['Min'], 'Max':value_dsf['Max']}
                                            edited_headers.append(header_to_edit)

                                            mandatory_edited_flag = True
                                        elif record[header_with_custom_rule] != value_dsf[j]:
                                            mandatory_edited_flag = False
                                else:

                                    if count_col != len(custom_rules_fields):

                                        cond_string = "Condition "+str(count_col)
                                        dsf_cond = value_dsf[cond_string]
                                    try:
                                        actual = record[value_dsf[i]]
                                        actual_counter_str = "Actual "+str(actual_counter)

                                        custom_rules_dict_temp = {j:value_dsf[j],actual_counter_str:actual}

                                        custom_rules_dict_temp.update(custom_rules_dict_temp)
                                        dict_lists_to_check.append(custom_rules_dict_temp)
                                    except KeyError:
                                        pass

                                    if actual_counter == field_counter:
                                        #
                                        if(Rules.check_condition_custom_rules(dict_lists_to_check,actual_conditions)):
                                            header_to_edit = value_dsf['DSF Field']
                                            if value_dsf['Field Attribute'] == 'P':
                                                if not header_to_edit in mandatory_p_temp:
                                                    mandatory_p_temp.append(header_to_edit)
                                                DSF_fields_p[header_to_edit] = {'Field Attribute':value_dsf['Field Attribute'],'Length':value_dsf['Length'],
                                                'Type':value_dsf['Type'],'Value':value_dsf['Consistent Default Value'], 'Min':value_dsf['Min'], 'Max':value_dsf['Max']}
                                            elif value_dsf['Field Attribute'] == 'C':
                                                if not header_to_edit in mandatory_c_temp:
                                                    mandatory_c_temp.append(header_to_edit)
                                                DSF_fields_c[header_to_edit] =  {'Field Attribute':value_dsf['Field Attribute'],'Length':value_dsf['Length'],
                                                'Type':value_dsf['Type'],'Value':value_dsf['Consistent Default Value'], 'Min':value_dsf['Min'], 'Max':value_dsf['Max']}
                                            elif value_dsf['Field Attribute'] =='B':
                                                if not header_to_edit in mandatory_p_temp:
                                                    mandatory_p_temp.append(header_to_edit)
                                                if not header_to_edit in mandatory_c_temp:
                                                    mandatory_c_temp.append(header_to_edit)
                                                DSF_fields_c[header_to_edit] =  {'Field Attribute':value_dsf['Field Attribute'],'Length':value_dsf['Length'],
                                                'Type':value_dsf['Type'],'Value':value_dsf['Consistent Default Value'], 'Min':value_dsf['Min'], 'Max':value_dsf['Max']}
                                                DSF_fields_p[header_to_edit] =  {'Field Attribute':value_dsf['Field Attribute'],'Length':value_dsf['Length'],
                                                'Type':value_dsf['Type'],'Value':value_dsf['Consistent Default Value'], 'Min':value_dsf['Min'], 'Max':value_dsf['Max']}
                                            else:
                                                if header_to_edit in mandatory_p_temp:
                                                    mandatory_p_temp.remove(header_to_edit)
                                                if header_to_edit in mandatory_c_temp:
                                                    mandatory_c_temp.remove(header_to_edit)
                                                DSF_fields_c[header_to_edit] =  {'Field Attribute':value_dsf['Field Attribute'],'Length':value_dsf['Length'],
                                                'Type':value_dsf['Type'],'Value':value_dsf['Consistent Default Value'], 'Min':value_dsf['Min'], 'Max':value_dsf['Max']}
                                                DSF_fields_p[header_to_edit] =  {'Field Attribute':value_dsf['Field Attribute'],'Length':value_dsf['Length'],
                                                'Type':value_dsf['Type'],'Value':value_dsf['Consistent Default Value'], 'Min':value_dsf['Min'], 'Max':value_dsf['Max']}
                                            mandatory_edited_flag = True
                                            edited_headers.append(header_to_edit)
                                    else:
                                        for d in dsf_fields:
                                            header_to_edit = value_dsf['DSF Field']
                                            if d['Field Attribute'] == 'P':
                                                DSF_fields_p[d['DSF FIELDS <START>']] ={'Field Attribute':d['Field Attribute'],'Length':d['Length'],
                                                'Type':d['Type'],'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX']}
                                            elif d['Field Attribute'] == 'C':
                                                # mandatory_c.append(d['DSF FIELDS <START>'])
                                                DSF_fields_c[d['DSF FIELDS <START>']] = {'Field Attribute':d['Field Attribute'],'Length':d['Length'],
                                                'Type':d['Type'],'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX']}
                                            elif d['Field Attribute'] =='B':
                                                # if header_to_edit in mandatory_p_temp:
                                                #     print("remove this shit")
                                                DSF_fields_c[d['DSF FIELDS <START>']] ={'Field Attribute':d['Field Attribute'],'Length':d['Length'],
                                                'Type':d['Type'],'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX']}
                                                DSF_fields_p[d['DSF FIELDS <START>']] = {'Field Attribute':d['Field Attribute'],'Length':d['Length'],
                                                'Type':d['Type'],'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX']}
                                            else:
                                                # print(d['DSF FIELDS <START>'])
                                                DSF_fields_c[d['DSF FIELDS <START>']] = {'Field Attribute':d['Field Attribute'],'Length':d['Length'],
                                                'Type':d['Type'],'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX']}
                                                DSF_fields_p[d['DSF FIELDS <START>']] = {'Field Attribute':d['Field Attribute'],'Length':d['Length'],
                                                'Type':d['Type'],'Value':d['Consistent Default Value'], 'Min':d['MIN'], 'Max':d['MAX']}
                                            mandatory_edited_flag = True
            if dsf_type == 1:
                record['POLICY NUMBER'] = policy_number_holder

            if client_type == 'P':
                if mandatory_edited_flag == True:

                    mis_p= Rules.check_mandatory(record, mandatory_p_temp)
                else:
                    mis_p= Rules.check_mandatory(record, mandatory_p)

                if mis_p: #check if missing data on a header
                    mf_h.append(mis_p)
                    mf_rows.append(record)
                    print("Missing: ",mis_p)
                try:
                  # if dsf_type==2 and record['Contract Type']:
                  if contract_type_field and validation_rules['Contract type validation']:


                      if not Rules.check_contract_type(record[contract_type_field].lower(), valid_contract_types):
                          t = record[contract_type_field] + " (Invalid Contract Type)"

                          if not record in invalid_tl:
                              invalid_tl.append(record)

                          if len(invalid_tl) != len(others_tl):
                              others_tl.append([])

                          if not record in others_rows:
                              others_rows.append(record)

                          others_tl[-1].append(t)
                except KeyError:
                  pass
                except:
                  if validation_rules['Contract type validation']:
                      print("Validator tried to process Contract Type")
                  else:
                      pass

                for header, req in DSF_fields_p.items():# Record level validation# 2

                    if dsf_type==1 and policy_type == 'DP':
                        if req['Min'] or req['Max']:

                            if not(Rules.check_range(record[header], req['Min'], req['Max'])):

                                l = header + "(DP Range)"
                                if not record in invalid_tl:
                                    invalid_tl.append(record)

                                if len(invalid_tl) != len(others_tl):
                                    others_tl.append([])

                                if not record in others_rows:
                                    others_rows.append(record)

                                others_tl[-1].append(l)

                    elif dsf_type==1 and policy_type == 'PC':
                        if req['Min'] or req['Max']:

                            if not(Rules.check_range(record[header], req['Min'], req['Max'])):

                                print(record[header],req['Min'],req['Max'])
                                l = header + "(PC Range)"
                                if not record in invalid_tl:
                                    invalid_tl.append(record)

                                if len(invalid_tl) != len(others_tl):
                                    others_tl.append([])

                                if not record in others_rows:
                                    others_rows.append(record)

                                others_tl[-1].append(l)

                    elif dsf_type==0:
                        if req['Min'] or req['Max']:
                            if not(Rules.check_range(record[header], req['Min'], req['Max'])):

                                l = header + "(Range)"
                                if not record in invalid_tl:
                                    invalid_tl.append(record)

                                if len(invalid_tl) != len(others_tl):
                                    others_tl.append([])

                                if not record in others_rows:
                                    others_rows.append(record)

                                others_tl[-1].append(l)

                    if not (Rules.check_length(record[header], req['Length'])):


                        l = header + "(Length)"
                        if not record in invalid_tl:
                            invalid_tl.append(record)

                        if len(invalid_tl) != len(others_tl):
                            others_tl.append([])

                        if not record in others_rows:
                            others_rows.append(record)

                        others_tl[-1].append(l)

                    if not Rules.check_type(record[header], req['Type']):

                        t = header + "(Type)"

                        if not record in invalid_tl:
                            invalid_tl.append(record)

                        if len(invalid_tl) != len(others_tl):
                            others_tl.append([])

                        if not record in others_rows:
                            others_rows.append(record)

                        others_tl[-1].append(t)

                if not record in invalid_tl:
                    valid_tl.append(record)


                for i in others_tl:
                    if len(others_tl) >=1:
                        if len(i)==0:
                            others_tl.pop()

            elif client_type == 'C':

                if mandatory_edited_flag == True:
                    mis_p= Rules.check_mandatory(record, mandatory_c_temp)

                else:

                    mis_p= Rules.check_mandatory(record, mandatory_c)

                if mis_p:
                    mf_h.append(mis_p)
                    mf_rows.append(record)


                if record[contract_type_field] and validation_rules['Contract type validation']:

                    if not Rules.check_contract_type(record[contract_type_field].lower(), valid_contract_types):
                        t = record['Contract Type'] + " (Invalid Contract Type)"

                        if not record in invalid_tl:
                            invalid_tl.append(record)

                        if len(invalid_tl) != len(others_tl):
                            others_tl.append([])

                        if not record in others_rows:
                            others_rows.append(record)

                        others_tl[-1].append(t)

                for header, req in DSF_fields_c.items():# Record level validation# 2
                    if dsf_type==1 and policy_type == 'DP':
                        if req['Min'] or req['Max']:

                            if not(Rules.check_range(record[header], req['Min'], req['Max'])):

                                l = header + "(DP Range)"

                                if not record in invalid_tl:
                                    invalid_tl.append(record)

                                if len(invalid_tl) != len(others_tl):
                                    others_tl.append([])

                                if not record in others_rows:
                                    others_rows.append(record)

                                others_tl[-1].append(l)

                    elif dsf_type==1 and policy_type == 'PC':
                        if req['Min'] or req['Max']:

                            if not(Rules.check_range(record[header], req['Min'], req['Max'])):

                                l = header + "(PC Range)"
                                if not record in invalid_tl:
                                    invalid_tl.append(record)

                                if len(invalid_tl) != len(others_tl):
                                    others_tl.append([])

                                if not record in others_rows:
                                    others_rows.append(record)

                                others_tl[-1].append(l)

                    elif dsf_type==0:
                        if req['Min'] or req['Max']:

                            if not(Rules.check_range(record[header], req['Min'], req['Max'])):

                                l = header + "(Range)"
                                if not record in invalid_tl:
                                    invalid_tl.append(record)

                                if len(invalid_tl) != len(others_tl):
                                    others_tl.append([])

                                if not record in others_rows:
                                    others_rows.append(record)

                                others_tl[-1].append(l)

                    if len(invalid_tl) != len(others_tl):
                            others_tl.append([])

                    if not (Rules.check_length(record[header], req['Length'])):

                        l = header + "(Length)"

                        if not record in invalid_tl:
                            invalid_tl.append(record)

                        if len(invalid_tl) != len(others_tl):
                            others_tl.append([])


                        if not record in others_rows:
                            others_rows.append(record)

                        others_tl[-1].append(l)


                    if not Rules.check_type(record[header], req['Type']):

                        t = header + "(Type)"

                        if not record in invalid_tl:
                            invalid_tl.append(record)

                        if len(invalid_tl) != len(others_tl):
                            others_tl.append([])


                        if not record in others_rows:
                            others_rows.append(record)

                        others_tl[-1].append(t)

                if not record in invalid_tl:
                    valid_tl.append(record)


                for i in others_tl:
                    if len(others_tl) >=1:
                        if len(i)==0:
                            others_tl.pop()


    date = datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')
    timestamp = datetime.strptime(date,'%m/%d/%Y %I:%M:%S %p').strftime('%Y%m%d%H%M%S')
    file_name = get_file_name(dsf)

    if dsf_type>=0 and dsf_type<=4:
        # print("invalid")
        inv = invalid_tl + mf_rows
        iv = []
        for i in range(len(inv)):
            if inv[i] not in inv[i + 1:]:
                iv.append(inv[i])

        delete_from_valid_tl = []


        for i in range(len(iv)):
            if iv[i] in valid_tl:
                delete_from_valid_tl.append(iv[i])

        new_valid = []
        new_valid = [x for x in valid_tl if x not in delete_from_valid_tl]

        print("Invalid:")
        for index in range(len(iv)):
            try:
              print(iv[index]['Name1'], iv[index]['Name2'])
            except KeyError:
                if dsf_type == 1:
                    print(iv[index]['POLICY NUMBER'])
                elif dsf_type ==4 or dsf_type==0:
                    print(iv[index]['Policy No.'])
                else:
                    print(iv[index]['Policy Number'])
            except:
              print("Policy number not found in DSF")


        print("Valid:")
        for index in range(len(new_valid)):
            try:
              print(new_valid[index]['Policy No'])
            except KeyError:
                if dsf_type == 1:
                    print(new_valid[index]['POLICY NUMBER'])
                elif dsf_type ==4:
                    print(new_valid[index]['Policy No.'])
                else:
                    print(new_valid[index]['Policy Number'])
            except:
              print("Policy number not found in DSF")

        if new_valid != []:
            write_to_csv(new_valid, f"{dir_name}\\{file_name} - VALID REC.csv")

        if invalid_tl != [] or mf_rows != []:
            inv = invalid_tl + mf_rows
            iv = []
            for i in range(len(inv)):
                if inv[i] not in inv[i + 1:]:
                    iv.append(inv[i])
            write_to_csv(iv, f"{dir_name}\\{file_name} - INVALID REC.csv")

    inv =[]

    if len(others_tl) == 0:
        print("")
    elif len(others_tl) >=1:
        if len(others_tl[0]) == 0:
            others_tl.pop(0)


    template = load_workbook(comp_template)

    missing_template = ""
    others_template = ""

    for i in template.sheetnames:
        if "MISSING" in i:
            missing_template = i
        if "OTHERS" in i:
            others_template = i

    ws_mf = template[missing_template]
    ws_others = template[others_template]
    comp_report_missing_temp = []
    comp_report_inv_temp = []

    for i in range(1, ws_mf.max_row+1):
        if i==5:
            for j in range(1, ws_mf.max_column+1):
                cell_obj = ws_mf.cell(row=i, column=j)
                comp_report_missing_temp.append(cell_obj.value)

    for i in range(1, ws_others.max_row+1):
        if i==5:
            for j in range(1, ws_others.max_column+1):
                cell_obj = ws_others.cell(row=i, column=j)
                comp_report_inv_temp.append(cell_obj.value)

    print("\nInvalid/Others:")
    for i,x in enumerate(others_rows):
        if dsf_type==0:
            add = []
            try:
                if x['CLIENT ADDRESS 1']:
                    add.append(x['CLIENT ADDRESS 1'])
                if x['CLIENT ADDRESS 2']:
                    add.append(x['CLIENT ADDRESS 2'])
                if x['CLIENT ADDRESS 3']:
                    add.append(x['CLIENT ADDRESS 3'])
                if x['CLIENT ADDRESS 4']:
                    add.append(x['CLIENT ADDRESS 4'])
                if x['CLIENT POSTCODE']:
                    add.append(x['CLIENT POSTCODE'])
                address = ', '.join(add)

                p = [x['POLICY NO'], '', x['FULL NAME (First Middle Last)'],'',
                x['CLIENT DATE OF BIRTH'], address, x['TAX IDENTIFICATION NUMBER'],
                '', ''  ,'','',date]

                p[-3] = ', '.join(others_tl[i])
                print(x['POLICY NO'],others_tl[i])
                inv.append(p)
            except KeyError:
                print("hello world")
                if x['Personal Street']:
                    add.append(x['Personal Street'])
                if x['Personal Line 1']:
                    add.append(x['Personal Line 1'])
                if x['Personal Line 2']:
                    add.append(x['Personal Line 2'])
                if x['Personal Line 3']:
                    add.append(x['Personal Line 3'])
                if x['Personal Post Code']:
                    add.append(x['Personal Post Code'])
                address = ', '.join(add)

                p = [x['Policy No.'], '', '', x['Personal Surname'], x['Personal Given Name'],'',
                x['Personal Birth Date'], address, x['Personal TIN #'],
                '', ''  ,'','','',date]

                p[-3] = ', '.join(others_tl[i])
                print(x['Policy No.'],others_tl[i])
                inv.append(p)
        elif dsf_type==4:
            add = []
            if x['Personal Line 1']:
                add.append(x['Personal Line 1'])
            if x['Personal Line 2']:
                add.append(x['Personal Line 2'])
            if x['Personal Line 3']:
                add.append(x['Personal Line 3'])
            if x['Personal Post Code']:
                add.append(x['Personal Post Code'])
            address = ', '.join(add)

            p = [x['Policy No.'], #policy no
            x['Client Type 1'], #client type
            x['Expiry Date'], #expiry date
            x['Personal Surname'], #Surname
            x['Personal Given Name'],#given name
            '',#corporate name
            x['Personal Birth Date'],#date of BIRTH
            address,
            x['Personal TIN #'], #TIN
            '',#registration no
            '', #screen header
            '',#screen no
            '', #exception message
            '', #screenshot
            date]#date

            p[-3] = ', '.join(others_tl[i])
            inv.append(p)
            print(x['Policy No.'],others_tl[i])
        elif dsf_type==1:
            add = []
            if x['CLIENT ADDRESS 1']:
                add.append(x['CLIENT ADDRESS 1'])
            if x['CLIENT ADDRESS 2']:
                add.append(x['CLIENT ADDRESS 2'])
            if x['CLIENT ADDRESS 3']:
                add.append(x['CLIENT ADDRESS 3'])
            if x['CLIENT ADDRESS 4']:
                add.append(x['CLIENT ADDRESS 4'])
            if x['CLIENT POSTCODE']:
                add.append(x['CLIENT POSTCODE'])
            address = ', '.join(add)

            p = [x['POLICY NUMBER'], '', x['FULL NAME (First, Middle, Last)'],'',
            x['CLIENT DATE OF BIRTH'], address, x['TAX IDENTIFICATION NUMBER'],
            '', ''  ,'','',date]


            p[-3] = ', '.join(others_tl[i])
            inv.append(p)
            print(x['POLICY NUMBER'],others_tl[i])


        elif dsf_type == 2 or dsf_type==3:
            add = []
            try:
              if x['Address Line 1']:
                  add.append(x['Address Line 1'])
              if x['Address Line 2']:
                  add.append(x['Address Line 2'])
              if x['Address Line 3']:
                  add.append(x['Address Line 3'])
              if x['Address Line 4']:
                  add.append(x['Address Line 4'])
              if x['Postal Code']:
                  add.append(x['Postal Code'])
              address = ', '.join(add)
            except KeyError:
              print("OSP Renewal - Change of Address Fields")
              if x['Permanent Address Line 1']:
                  add.append(x['Permanent Address Line 1'])
              if x['Permanent Address Line 2']:
                  add.append(x['Permanent Address Line 2'])
              if x['Permanent Address Line 3']:
                  add.append(x['Permanent Address Line 3'])
              if x['Permanent Address Line 4']:
                  add.append(x['Permanent Address Line 4'])
              if x['Permanent Postal Code']:
                  add.append(x['Permanent Postal Code'])
              address = ', '.join(add)
            except:
              print("Validator trying to process Address Fields")

            try:
                full_name = x['Name2'] + " " + x['Name2']
            except KeyError:
                full_name = x['Surname']
            except:
              print("Validator trying to append Full Name")


            try:
                p = [x['Policy No'], x['Client Type'], x['Expiry Date'], x['Name1'],x['Name2'],
                '',x['Date of Birth'], address, x['Tax Identification No'],
                x['Registration No'],'','','','',date,
                x['Agent Code'],x['Source Code'],
                x['Source Extension']]
            except KeyError:
                p = [x['Policy Number'], #policy number
                "Personal", #client type
                "",#Expiry
                x['Surname'],#Surname
                "",#First name
                "",#corporate NAME
                "",#date of birth
                address,#address
                "",#TIN
                "",#Registration number
                "",#Screen header
                "",#Screen no
                "",#invalid fields
                "",#screenshot
                date,#date
                "",#Agent
                x['Source Code'],#Source Code
                x['Source Ext'],#Source Extention
                x['BMA/AO']#BMA
                ]
            except:
              print("Validator trying to write in completion report")

            p[-5] = ', '.join(others_tl[i])
            inv.append(p)
            if dsf_type==3:
                try:
                    print(x['Policy Number'],others_tl[i])
                except KeyError:
                    print(x['Policy No'],others_tl[i])
            else:
                print(x['Policy No'],others_tl[i])

    mf = []
    print("\nMissing: ")
    for i,x in enumerate(mf_rows):
        add = []

        if dsf_type == 0:
            add = []
            try:
                if x['CLIENT ADDRESS 1']:
                    add.append(x['CLIENT ADDRESS 1'])
                if x['CLIENT ADDRESS 2']:
                    add.append(x['CLIENT ADDRESS 2'])
                if x['CLIENT ADDRESS 3']:
                    add.append(x['CLIENT ADDRESS 3'])
                if x['CLIENT ADDRESS 4']:
                    add.append(x['CLIENT ADDRESS 4'])
                if x['CLIENT POSTCODE']:
                    add.append(x['CLIENT POSTCODE'])
                address = ', '.join(add)

                p = [x['POLICY NO'], '', x['FULL NAME (First Middle Last)'],'',
                x['CLIENT DATE OF BIRTH'], address, x['TAX IDENTIFICATION NUMBER'],
                '',date]

                p[-2] = ', '.join(mf_h[i])
                mf.append(p)
                print(x['POLICY NO'],mf_h[i])
            except KeyError:
                # print("hello world - missing")
                if x['Personal Street']:
                    add.append(x['Personal Street'])
                if x['Personal Line 1']:
                    add.append(x['Personal Line 1'])
                if x['Personal Line 2']:
                    add.append(x['Personal Line 2'])
                if x['Personal Line 3']:
                    add.append(x['Personal Line 3'])
                if x['Personal Post Code']:
                    add.append(x['Personal Post Code'])
                address = ', '.join(add)

                p = [x['Policy No.'], '', '',x['Personal Surname'], x['Personal Given Name'],'',
                x['Personal Birth Date'], address, x['Personal TIN #'],
                '','',date]

                p[-2] = ', '.join(mf_h[i])
                mf.append(p)
                print(x['Policy No.'],mf_h[i])
        # elif client_type=='P' and dsf_type == 1:
        elif dsf_type==4:
            add = []
            if x['Personal Line 1']:
                add.append(x['Personal Line 1'])
            if x['Personal Line 2']:
                add.append(x['Personal Line 2'])
            if x['Personal Line 3']:
                add.append(x['Personal Line 3'])
            if x['Personal Post Code']:
                add.append(x['Personal Post Code'])
            address = ', '.join(add)

            p = [x['Policy No.'], #policy no
            x['Client Type 1'], #client type
            x['Expiry Date'], #expiry date
            x['Personal Surname'], #Surname
            x['Personal Given Name'],#given name
            '',#corporate name
            x['Personal Birth Date'],#date of BIRTH
            address,
            x['Personal TIN #'], #TIN
            '',#registration no
            '', #exception message
            date]#date


            p[-2] = ', '.join(mf_h[i])
            mf.append(p)
            print(x['Policy No.'],mf_h[i])
        elif dsf_type == 1:
            add = []
            if x['CLIENT ADDRESS 1']:
                add.append(x['CLIENT ADDRESS 1'])
            if x['CLIENT ADDRESS 2']:
                add.append(x['CLIENT ADDRESS 2'])
            if x['CLIENT ADDRESS 3']:
                add.append(x['CLIENT ADDRESS 3'])
            if x['CLIENT ADDRESS 4']:
                add.append(x['CLIENT ADDRESS 4'])
            if x['CLIENT POSTCODE']:
                add.append(x['CLIENT POSTCODE'])
            address = ', '.join(add)

            p = [x['POLICY NUMBER'], '', x['FULL NAME (First, Middle, Last)'],'',
            x['CLIENT DATE OF BIRTH'], address, '',
            '',date]

            p[-2] = ', '.join(mf_h[i])
            mf.append(p)
            print(x['POLICY NUMBER'],mf_h[i])

        elif dsf_type == 2 or dsf_type==3:
            add = []
            try:
              if x['Address Line 1']:
                  add.append(x['Address Line 1'])
              if x['Address Line 2']:
                  add.append(x['Address Line 2'])
              if x['Address Line 3']:
                  add.append(x['Address Line 3'])
              if x['Address Line 4']:
                  add.append(x['Address Line 4'])
              if x['Postal Code']:
                  add.append(x['Postal Code'])
              address = ', '.join(add)
            except KeyError:
              print("OSP Renewal - Change of Address Fields")
              if x['Permanent Address Line 1']:
                  add.append(x['Permanent Address Line 1'])
              if x['Permanent Address Line 2']:
                  add.append(x['Permanent Address Line 2'])
              if x['Permanent Address Line 3']:
                  add.append(x['Permanent Address Line 3'])
              if x['Permanent Address Line 4']:
                  add.append(x['Permanent Address Line 4'])
              if x['Permanent Postal Code']:
                  add.append(x['Permanent Postal Code'])
              address = ', '.join(add)
            except:
              print("Validator trying to process Address Fields")

            try:
                full_name = x['Name2'] + " " + x['Name2']
            except KeyError:
                full_name = x['Surname']
            except:
              print("Validator trying to append Full Name")

            try:
                p = [x['Policy No'], x['Client Type'], x['Expiry Date'], x['Name1'],x['Name2'],
                '',x['Date of Birth'], address, x['Tax Identification No'],
                x['Registration No'],x['Agent Code'],x['Source Code'],
                x['Source Extension'],'',date]
            except KeyError:
                p = [x['Policy Number'], #policy number
                "Personal", #client type
                "",#Expiry
                x['Surname'],#Surname
                "",#First name
                "",#corporate NAME
                "",#date of birth
                address,#address
                "",#TIN
                "",#Registration number
                "",#Agent
                x['Source Code'],#Source Code
                x['Source Ext'],#Source Extention
                "",#missing fields
                date,#date
                x['BMA/AO']#BMA
                ]
            except:
              print("Validator trying to write in completion report")

            p[-2] = ', '.join(mf_h[i])
            mf.append(p)
            try:
                print(x['Policy Number'],mf_h[i])
            except KeyError:
                print(x['Policy No'],mf_h[i])


    ws_mf = template[missing_template]
    ws_mf.cell(row=1,column=2, value=date.split()[0])
    ws_mf.cell(row=2,column=2, value=' '.join(date.split()[1:]))
    ws_mf.cell(row=3,column=2, value=os.path.basename(dsf))

    for row in mf:
        ws_mf.append(row)

    Template.remove_emp_rows(ws_mf)

    ws_others = template[others_template]
    ws_others.cell(row=1,column=2, value=date.split()[0])
    ws_others.cell(row=2,column=2, value=' '.join((date.split()[1:])))
    ws_others.cell(row=3,column=2, value=os.path.basename(dsf))

    for row in inv:
        ws_others.append(row)

    Template.remove_emp_rows(ws_others)

    template.save(f"{dir_name}\\{file_name} - Completion Report.xlsx")

def write_to_xlsx(input_file,records,type,output_dir):

    output_file = input_file.split('\\')
    new_name = output_file[-1].split('.')[0]
    new_name = new_name + ' - ' + type +'.xlsx'

    output_file[-1] = new_name
    output_dir = output_dir + '\\' + new_name

    # output_file = '\\'.join(output_file)
    # print(records)
    # print(type(records))
    for i in records.items():
        # list_temp = i[1:]
        # print(list_temp)
        df = pd.DataFrame(i[1])
        # print(df)
        try:
            with pd.ExcelWriter(output_dir,engine='openpyxl', mode='a',  if_sheet_exists='overlay') as writer:
                df.to_excel(writer, sheet_name=i[0], index=False,header=False)
        except:
            with pd.ExcelWriter(output_dir,engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=i[0], index=False,header=False)
        finally:
            with pd.ExcelWriter(output_dir,engine='openpyxl', mode='a',  if_sheet_exists='overlay') as writer:
                df.to_excel(writer, sheet_name=i[0], index=False,header=False)
        # for key,value in i.items():
        #     print(key,value)
        #



def write_exception_list(input_file,output_file,file_format_exception,sheets_exception,headers_exception,at_least_one_record_exception):
    # print('******************')
    file_name = input_file.split('\\')[-1]
    file_name = file_name.split('.')[0]
    file_name = file_name + ' - Exception List.xlsx'
    output_file = output_file + '\\' + file_name

    sheets_exception_set = list(sheets_exception for sheets_exception,_ in itertools.groupby(sheets_exception))
    header_exception_set = list(headers_exception for headers_exception,_ in itertools.groupby(headers_exception))
    at_least_one_record_exception_set = list(at_least_one_record_exception for at_least_one_record_exception,_ in itertools.groupby(at_least_one_record_exception))

    df_file_format = pd.DataFrame(file_format_exception)
    df_sheets = pd.DataFrame(sheets_exception_set)
    df_headers = pd.DataFrame(header_exception_set)
    df_num_records = pd.DataFrame(at_least_one_record_exception_set)

    with pd.ExcelWriter(output_file) as writer:
        df_file_format.to_excel(writer, sheet_name='File Format', index=None, header=None)
        df_sheets.to_excel(writer, sheet_name='Sheets', index=None, header=None)
        df_headers.to_excel(writer, sheet_name='Headers', index=None, header=None)
        df_num_records.to_excel(writer, sheet_name='Number of Records', index=None, header=None)


def validator(config_file,input_file,sheet_names,output_dir):
    df_fields = pd.read_excel(config_file, sheet_name='Fields')
    headers_config_dict = df_fields.to_dict('records')
    list_of_invalid_records = []
    list_of_valid_records = []

    bool_length = True
    bool_type = True
    invalid_dict_temp = {}
    valid_dict_temp = {}

    for i in sheet_names:
        df_temp = pd.read_excel(input_file, sheet_name=i)
        records_temp = df_temp.to_dict('records')
        if len(records_temp) > 0:
            list_of_invalid_dicts = []
            list_of_valid_dicts = []

            headers = list(df_temp.columns)
            headers_invalid = list(df_temp.columns)
            headers_invalid.append('Remarks')

            new_valid_records = []

            valid_records = []
            valid_records.append(headers)
            invalid_records = []
            invalid_records.append(headers_invalid)

            for j in records_temp:
                j['Remarks'] = ''
                remarks = ''

                for key,value in j.items():
                    for m in headers_config_dict:
                        for key_m,value_m in m.items():
                            if i == m['Sheet']:
                                if key == m['Fields']:
                                    temp_list = []
                                    if key_m == 'Length':
                                        if not (Rules.check_length(str(j[key]), value_m)):

                                            remarks = remarks + 'Field <'+str(key)+'> has invalid length.\n'
                                            j['Remarks'] = remarks
                                            bool_length == False

                                        # elif bool_length == True and bool_type == True and j['Remarks'] == '':
                                            # temp_list = []
                                            # temp_list = list(j.values()).copy()
                                            # valid_records.append(j.values())


                                    if key_m == 'Type':
                                        # print(key_m)
                                        # print(len(str(i[key])),value_j)
                                        if not (Rules.check_type(str(j[key]), value_m)):
                                            remarks = remarks + 'Field <'+str(key)+'> has invalid data type.\n'
                                            j['Remarks'] = remarks
                                            bool_type == False
                                        # elif bool_type == True and bool_length == True and j['Remarks'] == '':
                                            # temp_list = []
                                            # temp_list = list(j.values()).copy()
                                            # valid_records.append(j.values())

                # temp_list = list(j.values()).copy()  #do not indent
                temp_list = j.values()
                if j['Remarks'] == '' and bool_type == True and bool_length == True:  #do not indent
                    valid_records.append(list(j.values()))
                else:
                    # print(temp_list)
                    invalid_records.append(list(j.values()))#do not indent

                for elem in valid_records:
                    if elem not in new_valid_records:
                        new_valid_records.append(elem)
                valid_records = new_valid_records

                # print(invalid_records) # do not indent
                # print(valid_records)

                list_of_valid_dicts.append(valid_records)
                # print(invalid_records)
                list_of_invalid_dicts.append(invalid_records)

            # print(i,list_of_invalid_dicts)
                # print(invalid_records)
            dict_with_sheet_name_invalid = {i:invalid_records}
            dict_with_sheet_name_valid = {i:valid_records}
            # print(dict_with_sheet_name_invalid)
            #     # print(i,list_of_valid_dicts)
            #


            # temp_new = (i,list_of_invalid_dicts)
            write_to_xlsx(input_file,dict_with_sheet_name_invalid,'INVALID RECORDS',output_dir)
            write_to_xlsx(input_file,dict_with_sheet_name_valid,'VALID RECORDS',output_dir)


def read_config(config_file,input_file,output_dir):
    input_file_pd = pd.ExcelFile(input_file)

    df_rules = pd.read_excel(config_file, sheet_name='Project Info')
    rules_list = df_rules.values.tolist()

    df_sheets = pd.read_excel(config_file, sheet_name='Valid sheets')
    valid_sheet_names = df_sheets.values.tolist()

    df_fields = pd.read_excel(config_file, sheet_name='Fields')
    headers_config_dict = df_fields.to_dict('records')

    df_records = pd.read_excel(input_file, sheet_name='Worksheet')
    df_records_all = pd.ExcelFile(input_file)
    records_list = df_records.values.tolist()
    headers = list(df_records.columns)
    # print(headers)

    rule_name = []
    rule_value = []

    # headers = []
    records = []

    #List of Exceptions
    file_format_exception = []
    file_format_exception_headers = ['Filename','Remarks']
    file_format_exception.append(file_format_exception_headers)

    sheets_exception = []
    sheets_exception_headers = ['Missing Sheets']
    sheets_exception.append('Missing Sheets')

    headers_exception = []
    headers_exception_headers = ['Sheet Name','Missing Fields']
    headers_exception.append(headers_exception_headers)

    at_least_one_record_exception = []
    at_least_one_record_exception_headers = ['Remarks']
    at_least_one_record_exception.append(at_least_one_record_exception_headers)

    temp_dict = {}


    for i in rules_list:
        rule_name.append(i[0])
        rule_value.append(i[1])

    validation_rules = dict(zip(rule_name,rule_value))
    if validation_rules['Validate file is XLSX']:
        if not Rules.check_xlsx(input_file):
            temp_exception = [input_file,'INVALID']
            file_format_exception.append(temp_exception)
            write_exception_list(input_file,output_dir,file_format_exception,sheets_exception,headers_exception,at_least_one_record_exception)
            raise Exception('Input file not an XLSX file. Check file format.')

        else:
            temp_exception = [input_file,'VALID']
            file_format_exception.append(temp_exception)
            write_exception_list(input_file,output_dir,file_format_exception,sheets_exception,headers_exception,at_least_one_record_exception)
            # print('Input file is an XLSX file. Correct file format.')

    #Continue, append to exception file
    pass_sheet_names = []
    if validation_rules['Sheet validation']:
        for i in valid_sheet_names:
            if not i[0] in input_file_pd.sheet_names:
                exception = 'Sheet ' + i[0] + ' not found in input file.'
                sheets_exception.append(i[0])

            else:
                pass_sheet_names.append(i[0])
        write_exception_list(input_file,output_dir,file_format_exception,sheets_exception,headers_exception,at_least_one_record_exception)

        # for i in pass_sheet_names:
        #     print(i)

    #Continue, append to exception file
    headers_config_list = []

    for i in headers_config_dict:
        # print(i)
        headers_config_list.append(i['Fields'])

    sheet_names = df_records_all.sheet_names
    if validation_rules['Validate Header Fields']:
        for sheet in sheet_names:
            missing_headers = []

            df_temp = pd.read_excel(input_file, sheet_name=sheet)
            headers_temp = list(df_temp.columns)

            if sheet in pass_sheet_names:
                for j in headers_config_dict:
                    # print(j)
                    if sheet == j['Sheet']:
                        # print(sheet,j['Sheet'])
                        if not j['Fields'] in headers_temp:
                            missing_headers.append(j['Fields'])
                            temp_dict = {'Sheet Name': sheet,'Missing Headers':missing_headers}
                            headers_exception.append(list(temp_dict.values()))
        # print(headers_exception)
        write_exception_list(input_file,output_dir,file_format_exception,sheets_exception,headers_exception,at_least_one_record_exception)


    # if validation_rules['At least one record'] and not Rules.at_least_one_record(len(records_list)):
    #     raise Exception('File has no records.')

    if validation_rules['At least one record']:
        for sheet in sheet_names:
            df_temp = pd.read_excel(input_file, sheet_name=sheet)
            records_list = df_temp.values.tolist()

            if not Rules.at_least_one_record(len(records_list)):
                temp_list_string = []
                temp_string = sheet + ' sheet has no records.'
                temp_list_string.append(temp_string)
                at_least_one_record_exception.append(temp_list_string)
                # print(temp_string)

        write_exception_list(input_file,output_dir,file_format_exception,sheets_exception,headers_exception,at_least_one_record_exception)

    validator(config_file,input_file,pass_sheet_names,output_dir)


def validator_controller(*params):

    if len(params)==3:
        # print('\n\nGeneric validator activated')

        try:
            input_file = params[0]
            config_file = params[1]
            output_dir = params[2]

            print('\nInput: ',input_file)
            print('Config: ',config_file)
            print('Output: ',output_dir)
            print('\n\n')
        except Exception as e:
            print(e)

        current_directory = os.path.abspath(os.getcwd())
        for cd in os.listdir(current_directory):
            if input_file in cd:
                input_file = cd

        read_config(config_file,input_file,output_dir)

        #Input file
        #dev worksheet / config
        #Completion report
        #Output directory

        #1 Read config
        #2 Read Records
        #3 Validate Records


    if len(params)==5:
        parser = argparse.ArgumentParser()
        parser.add_argument("dsf")
        parser.add_argument("dev_worksheet")
        parser.add_argument("comp_template")
        parser.add_argument("arg_dsf_type")
        parser.add_argument("output_directory", nargs='?', default="")
        args =  parser.parse_args()
        policy_type = args.arg_dsf_type
        dsf_type = 0

        file_name = get_file_name(args.dsf)
        file_name_lower = file_name.lower()

        file_ext = args.dsf.split(".")[1]

        dir_name = os.path.join(args.output_directory, file_name)
        os.makedirs(dir_name, exist_ok=True)

        df_policy_types = pd.read_excel(args.dev_worksheet, sheet_name='Policy Types', usecols="A:D")
        # print(df_policy_types)
        dict_policy_types = df_policy_types.to_dict('records')

        default_policy_type = dict_policy_types[0]['default']
        print("Default Policy Type:", default_policy_type)
        # print(policy_type)

        for i in dict_policy_types:
            # print(i['user_input'])
            # print(policy_type)
            if i['user_input'] == policy_type:
                print("User input exists in config - ", i['user_input'])
                if i['file_name'] in args.dsf or (not i['file_name'] in args.dsf):

                    try:
                        if policy_type=="osp":
                            dsf_type = 2
                        elif policy_type=="motor":
                            dsf_type = 1
                        elif policy_type=="pa":
                            dsf_type = 0
                        elif policy_type=="osprenewal":
                            dsf_type = 3
                        elif policy_type=="webapppa":
                            dsf_type = 4
                        else:
                            raise Exception('Check naming convention of dsf file. Should contain "motor", "pa", or "osp".')
                        dsf_validator(args.dsf, args.dev_worksheet, args.comp_template, dsf_type, dir_name, file_ext)
                    except Exception as exception:
                        traceback.print_exception(
                            type(exception),
                            exception,
                            exception.__traceback__
                        )
                        with open(f"{dir_name}\\{file_name} - error.log", "w") as file:
                            file.write(f"Error in file {args.dsf}: {exception}")
                else:
                    try:
                        if default_policy_type=="osp":
                            dsf_type = 2
                        elif default_policy_type=="motor":
                            dsf_type = 1
                        elif default_policy_type=="pa":
                            dsf_type = 0
                        elif default_policy_type=="osprenewal":
                            dsf_type = 3
                        elif default_policy_type=="webapppa":
                            dsf_type = 4
                        dsf_validator(args.dsf, args.dev_worksheet, args.comp_template, dsf_type, dir_name, file_ext)
                    except Exception as exception:
                        traceback.print_exception(
                            type(exception),
                            exception,
                            exception.__traceback__
                        )
                        with open(f"{dir_name}\\{file_name} - error.log", "w") as file:
                            file.write(f"Error in file {args.dsf}: {exception}")


def main(args=None):
    # try:
    validator_controller(*sys.argv[1:])
    # except Exception as e:
    #     print(e)


if __name__ == '__main__':
    start_time = time.time()
    main()
    print("--- %s seconds ---" % (time.time() - start_time))
